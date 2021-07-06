# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
import openpyxl as ox
import pandas as pd
from datetime import datetime

excel_file = '../database/testOld.xlsx'
op_file = "../database/testOld"+datetime.now().strftime("%m%d%Y_%H%M%S")+".xlsx"
wb = ox.load_workbook(excel_file)
ipSheet = wb['testOld']
opSheet = wb['output']


# %%
delim = ", "
rowOut=2
for rowIn in range(2,ipSheet.max_row+1):
    lncRNA = ipSheet.cell(rowIn,1).value
    num_transcripts = ipSheet.cell(rowIn,2).value
    if(num_transcripts>1):
        total_pqs = ipSheet.cell(rowIn,3).value.split(delim)
        g2 = ipSheet.cell(rowIn,4).value.split(delim)
        g3 = ipSheet.cell(rowIn,5).value.split(delim)
        g4 = ipSheet.cell(rowIn,6).value.split(delim)
        ncbi_id = ipSheet.cell(rowIn,7).value.split(delim)
    
    for i in range(len(total_pqs)):
        opSheet.cell(rowOut,1).value = lncRNA
        opSheet.cell(rowOut,2).value = i+1
        opSheet.cell(rowOut,3).value= int(total_pqs[i])
        opSheet.cell(rowOut,4).value= int(g2[i])
        opSheet.cell(rowOut,5).value= int(g3[i])
        opSheet.cell(rowOut,6).value= int(g4[i])
        opSheet.cell(rowOut,7).value= ncbi_id[i]
        rowOut+=1


# %%
wb.save(op_file)


